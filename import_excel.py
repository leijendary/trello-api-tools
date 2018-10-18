import configparser
import requests
from openpyxl import load_workbook

SUPP_DOCU_START = 'Supporting Documents:\n\n'
CLG_NOTES_START = 'Investigation Notes - CLG Systems:\n\n'
SP_NOTES_START = 'Investigation Notes - Service Provider:\n\n'

# load the configuration
config = configparser.ConfigParser()
config.read('config.cfg')

api_config = config['API']
board_config = config['Board']
card_config = config['Card']
action_config = config['Action']
list_config = config['List']
input_config = config['Input']

API_KEY = api_config.get('Key')
API_TOKEN = api_config.get('Token')

BOARD_ID = board_config.get('Id')
FILE_PATH = input_config.get('FilePath')

# the only fields that we will get from the cards api
CARD_FIELDS = card_config.get('Fields')

# build the cards url
CARD_LIST_URL = 'https://api.trello.com/1/boards/{}/cards?' \
    'fields={}&key={}&token={}'.format(BOARD_ID, CARD_FIELDS, API_KEY,
        API_TOKEN)

CARD_URL = 'https://api.trello.com/1/cards'
COMMENT_URL = 'https://api.trello.com/1/cards/{}/actions/comments'

# build the labels url
LABEL_LIST_URL = 'https://api.trello.com/1/boards/{}/labels?key={}&token={}' \
    .format(BOARD_ID, API_KEY, API_TOKEN)

# the only fields that we will get from the actions api
ACTION_FIELDS = action_config.get('Fields')
# the only data that we will get from the actions api
ACTION_FILTER = action_config.get('Filter')
# show the member creator field
ACTION_SHOW_MEMBER_CREATOR = action_config.get('MemberCreator')

# build the actions url
ACTION_LIST_URL = 'https://api.trello.com/1/cards/{}/actions?key={}&token={}' \
    '&fields={}&filter={}&memberCreator={}'

# backlog ID
BACKLOG_LIST_ID = list_config.get('BacklogId')
# closed ID
CLOSED_LIST_ID = list_config.get('ClosedId')

# get the cards from the url
cards = requests.get(CARD_LIST_URL).json()

# get the labels from the url
labels = requests.get(LABEL_LIST_URL).json()

# read the file to be imported to trello
# load workbook
wb = load_workbook(filename=FILE_PATH, read_only=True)
# get the issue log details worksheet
ws = wb[input_config.get('WorkbookName')]

# get the maximum rows
MAX_ROWS = ws.max_row

# read each row on the file
def read_rows():
    # start at row 5.
    row_number = 4

    # report counters
    closed_count = 0
    reopen_count = 0
    existing_count = 0
    new_count = 0
    new_comment_count = 0

    for row in ws.iter_rows(row_offset=4, max_row=MAX_ROWS):
        # increment row number for the next issue
        row_number += 1

        # get the IR number
        ir = row[9].value
        status = row[11].value

        # check if the status is not closed. if the status is closed, skip this
        if status.lower() == 'closed':
            print(ir + ' is closed')

            card = get_card_by_ir(ir);

            if card is not None:
                if card['idList'] != CLOSED_LIST_ID:
                    move_card_list(card['id'], CLOSED_LIST_ID)

                    # update comments of the card if there are any updates
                    new_comment_count += update_comments(card['id'], row)

                    print(ir + ' is moved to closed')
            else:
                create_card(row_number, ir, row, CLOSED_LIST_ID)

                print(ir + ' is created in the closed list')

            closed_count += 1

            continue
        elif status.lower() == 're-open':
            print(ir + ' is re-opened')

            card = get_card_by_ir(ir);

            if card is not None:
                if card['idList'] != BACKLOG_LIST_ID:
                    move_card_list(card['id'], BACKLOG_LIST_ID)

                    # update comments of the card if there are any updates
                    new_comment_count += update_comments(card['id'], row)

                    print(ir + ' is moved to backlog')
            else:
                create_card(row_number, ir, row, BACKLOG_LIST_ID)

                print(ir + ' is created in the backlog list')

            reopen_count += 1

        # check if the IR number is already existing
        if has_ir_already(ir):
            card = get_card_by_ir(ir);

            # update comments of the card if there are any updates
            new_comment_count += update_comments(card['id'], row)

            print(ir + ' already exists')

            existing_count += 1

            continue

        # create the card in the API
        create_card(row_number, ir, row, BACKLOG_LIST_ID)

        new_count += 1

    current = new_count + existing_count
    total = closed_count + reopen_count + current

    print('Closed : {}'.format(closed_count))
    print('Re-Opened : {}'.format(reopen_count))
    print('Current: {}'.format(current))
    print('New : {}'.format(new_count))
    print('Total : {}'.format(total))
    print('Additional Comments : {}'.format(new_comment_count))


# create a card based on the ir and row
def create_card(row_num, ir, row, list_id):
    # get the column values
    module = row[3].value
    problem_statement = row[4].value
    severity = row[6].value

    # build the title. the format is
    # "{ir number}: {first line of the problem statement}"
    title = '{}: {}'.format(ir, problem_statement.split('\n', 1)[0])

    # build the description of the card
    description = '{}\n\n' \
        '==========================================================\n\n' \
        '**Module:** {}\n\n' \
        '**Line:** #{}'.format(problem_statement, module, str(row_num))

    # get the label using the severity
    label_id = ''

    for label in labels:
        name = label['name']

        if name.lower() == severity.lower():
            label_id = label['id']

            break

    # create the card object
    param = {
        'name': title,
        'desc': description,
        'pos': 'bottom',
        'idList': list_id,
        'idLabels': label_id,
        'key': API_KEY,
        'token': API_TOKEN
    }

    # post the card to the API
    card = requests.post(CARD_URL, params=param).json()

    # add comments if there are any
    supp_docu = get_supp_docu(row)
    clg_notes = get_clg_notes(row)
    sp_notes = get_sp_notes(row)

    if supp_docu:
        create_comment(card['id'], SUPP_DOCU_START + str(supp_docu.strip()))

    if clg_notes:
        create_comment(card['id'], CLG_NOTES_START + str(clg_notes.strip()))

    if sp_notes:
        create_comment(card['id'], SP_NOTES_START + str(sp_notes.strip()))


# move the card to another list
def move_card_list(card_id, list_id):
    param = {
        'idList': list_id,
        'pos': 'bottom',
        'key': API_KEY,
        'token': API_TOKEN
    }

    requests.put(CARD_URL + '/{}'.format(card_id), params=param)


# update the comments section of the card
def update_comments(card_id, row):
    url = ACTION_LIST_URL.format(
        card_id,
        API_KEY,
        API_TOKEN,
        ACTION_FIELDS,
        ACTION_FILTER,
        ACTION_SHOW_MEMBER_CREATOR)
    # get all the comments from the card
    actions = requests.get(url).json()
    comment_count = 0

    # get the comments of the row
    supp_docu = get_supp_docu(row)
    clg_notes = get_clg_notes(row)
    sp_notes = get_sp_notes(row)

    # for each comment, check if the
    for action in actions:
        # if the three comments from the row are all not valid anymore,
        # exit this function
        if not supp_docu and not clg_notes and not sp_notes:
            return 0

        text = action['data']['text']

        # remove supporting document start title
        if text.startswith(SUPP_DOCU_START):
            text = text.replace(SUPP_DOCU_START, '')

            if supp_docu:
                supp_docu = supp_docu.replace(text, '')

        # remove clg notes start title
        if text.startswith(CLG_NOTES_START):
            text = text.replace(CLG_NOTES_START, '')

            if clg_notes:
                clg_notes = clg_notes.replace(text, '')

        # remove sp notes start title
        if text.startswith(SP_NOTES_START):
            text = text.replace(SP_NOTES_START, '')

            if sp_notes:
                sp_notes = sp_notes.replace(text, '')

        # if the supporting document starts with the action text, remove
        # the supporting document from the comments to be added
        if supp_docu:
            if supp_docu.lower().startswith(text.lower()):
                supp_docu = None

        # if the clg notes starts with the action text, remove the clg notes
        # from the comments to be added
        if clg_notes:
            if clg_notes.lower().startswith(text.lower()):
                clg_notes = None

        # if the sp notes starts with the action text, remove the sp notes
        # from the comments to be added
        if sp_notes:
            if sp_notes.lower().startswith(text.lower()):
                sp_notes = None

    if supp_docu:
        create_comment(card_id, SUPP_DOCU_START + str(supp_docu))

        comment_count += 1

    if clg_notes:
        create_comment(card_id, CLG_NOTES_START + str(clg_notes))

        comment_count += 1

    if sp_notes:
        create_comment(card_id, SP_NOTES_START + str(sp_notes))

        comment_count += 1

    return comment_count


# create a comment for the card
def create_comment(card_id, text):
    comment = {
        'text': text,
        'key': API_KEY,
        'token': API_TOKEN
    }

    requests.post(COMMENT_URL.format(card_id), comment)


# get the card by it's IR number
def get_card_by_ir(ir):
    for card in cards:
        name = card['name']

        if name.lower().startswith(ir.lower()):
            return card

    return None


# check if the list of cards has the IR number already
def has_ir_already(ir):
    return get_card_by_ir(ir) is not None


def get_supp_docu(row):
    return row[5].value


def get_clg_notes(row):
    return row[20].value


def get_sp_notes(row):
    return row[21].value


if __name__ == '__main__':
    read_rows()